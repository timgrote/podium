#!/usr/bin/env python3
"""Extract and import TIE invoice data from XLSX and PDF files.

Scans each imported project's dropbox_path for invoice files (XLSX preferred,
PDF fallback), parses both invoice formats (professional task-based + simple
T&M), and imports invoices WITH line items into the database.

Two invoice formats:
  Professional (XLSX sheet "Invoice"):
    Row 17: headers (Task, Fee, % Complete, Previous Fee Billing, Current Fee Billing)
    Rows 18-27: task data in columns B-F
    Cell C11: invoice number, C10: date, C9: project ID, F13: client project #

  Simple (XLSX sheet "Simple Invoice"):
    Row 21: headers (Description, Quantity, Rate, Amount)
    Rows 22-31: line items in columns A-D
    Cell D8: invoice number, D7: date, D9: description

Run modes:
  python3 scripts/extract_tie_invoices.py              # Preview only
  python3 scripts/extract_tie_invoices.py --import      # Import all
  python3 scripts/extract_tie_invoices.py --import-high  # Import 90%+ confidence only
"""

import json
import os
import re
import sys
import uuid
from datetime import datetime

import openpyxl
import psycopg2
import psycopg2.extras
import requests

API_BASE = "http://localhost:3000/api"
TIE_BASE = "/mnt/d/Dropbox/TIE"
DB_URL = os.environ.get(
    "CONDUCTOR_DATABASE_URL",
    "postgresql://conductor:conductor@localhost:5432/conductor",
)


def generate_id(prefix):
    return prefix + uuid.uuid4().hex[:8]


# ── File discovery ────────────────────────────────────────────────────

def find_invoice_files(data_path):
    """Find invoice XLSX and PDF files under a project's data_path.

    Returns dict: {base_name: {"xlsx": path, "pdf": path}}
    Groups matching XLSX/PDF pairs by stripping extension.
    """
    files = {"xlsx": [], "pdf": []}
    for root, dirs, filenames in os.walk(data_path):
        for f in filenames:
            lower = f.lower()
            if "invoice" not in lower:
                continue
            full = os.path.join(root, f)
            if lower.endswith(".xlsx"):
                files["xlsx"].append(full)
            elif lower.endswith(".pdf"):
                files["pdf"].append(full)
    return files


# ── XLSX parsing (primary, high confidence) ───────────────────────────

def parse_xlsx_professional(ws):
    """Parse professional services format XLSX.

    Layout: B17=headers, B18:F27=task rows
    B=Task name, C=Fee, D=% Complete, E=Previous Billing, F=Current Billing
    """
    result = {
        "format": "professional",
        "confidence": 0.95,
        "line_items": [],
    }

    # Header fields
    result["invoice_number"] = str(ws["C11"].value or "")
    date_val = ws["C10"].value
    if isinstance(date_val, datetime):
        result["date"] = date_val.strftime("%Y-%m-%d")
    elif date_val:
        result["date"] = str(date_val)
    result["project_ref"] = str(ws["C9"].value or "")
    result["client_project_number"] = str(ws["F13"].value or "")
    result["pm"] = str(ws["C12"].value or "")
    result["bill_to"] = str(ws["F9"].value or "")

    # Task rows: B18 through B27 (up to 10 tasks)
    for row_num in range(18, 28):
        name = ws[f"B{row_num}"].value
        fee = ws[f"C{row_num}"].value
        if not name or str(name).strip() == "Total":
            continue
        pct = ws[f"D{row_num}"].value
        prev = ws[f"E{row_num}"].value
        current = ws[f"F{row_num}"].value

        # Safely convert to float, treating non-numeric as 0
        def safe_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0.0

        result["line_items"].append({
            "name": str(name).strip(),
            "unit_price": safe_float(fee),
            "quantity": safe_float(pct),
            "previous_billing": safe_float(prev),
            "amount": safe_float(current),
        })

    # Invoice total from F30
    invoice_total = ws["F30"].value
    if invoice_total is not None:
        result["total_due"] = float(invoice_total)
    else:
        result["total_due"] = sum(li["amount"] for li in result["line_items"])

    # Contract total from C28
    contract_total = ws["C28"].value
    if contract_total is not None:
        result["contract_total"] = float(contract_total)

    if not result["invoice_number"]:
        result["confidence"] -= 0.2
    if not result["line_items"]:
        result["confidence"] -= 0.3

    return result


def parse_xlsx_simple(ws):
    """Parse simple invoice format XLSX.

    Layout: A21=headers, A22:D31=line items
    A=Description, B=Quantity, C=Rate, D=Amount
    Previous invoices listed in A14:D17
    """
    result = {
        "format": "simple",
        "confidence": 0.90,
        "line_items": [],
    }

    # Header fields
    result["invoice_number"] = str(ws["D8"].value or "")
    date_val = ws["D7"].value
    if isinstance(date_val, datetime):
        result["date"] = date_val.strftime("%Y-%m-%d")
    elif date_val:
        result["date"] = str(date_val)
    result["description"] = str(ws["D9"].value or "")
    result["bill_to"] = str(ws["A7"].value or "")
    # Client project number can be in D11 or D13 depending on layout
    cpn = ws["D11"].value
    if not cpn or (isinstance(cpn, str) and cpn.startswith("#")):
        cpn = ws["D13"].value
    if cpn:
        result["client_project_number"] = str(cpn).lstrip("#")

    # Previous invoice amounts (rows 14-17 area)
    prev_total = 0
    for row_num in range(14, 18):
        prev_inv = ws[f"A{row_num}"].value
        prev_amt = ws[f"D{row_num}"].value
        if prev_inv and prev_amt and isinstance(prev_amt, (int, float)):
            prev_total += float(prev_amt)

    # Line items: scan rows 21-32 for DESCRIPTION header then read items below it
    task_group = None
    start_row = 22
    # Find the header row dynamically
    for check_row in range(20, 25):
        val = ws[f"A{check_row}"].value
        if val and "DESCRIPTION" in str(val).upper():
            start_row = check_row + 1
            break
    for row_num in range(start_row, start_row + 10):
        name = ws[f"A{row_num}"].value
        if not name:
            continue
        qty = ws[f"B{row_num}"].value
        rate = ws[f"C{row_num}"].value
        amount = ws[f"D{row_num}"].value

        # If no qty/rate/amount, it's a task group header
        if qty is None and rate is None and amount is None:
            task_group = str(name).strip()
            continue

        if amount is not None and float(amount) > 0:
            display_name = str(name).strip()
            if task_group:
                display_name = f"{task_group} - {display_name}"
            result["line_items"].append({
                "name": display_name,
                "quantity": float(qty or 1),
                "unit_price": float(rate or 0),
                "amount": float(amount),
                "previous_billing": 0,  # simple format tracks prev separately
            })

    # Total due from D32
    total_due = ws["D32"].value
    if total_due is not None:
        result["total_due"] = float(total_due)
    else:
        result["total_due"] = sum(li["amount"] for li in result["line_items"])

    # Project total from D18
    project_total = ws["D18"].value
    if project_total is not None:
        result["contract_total"] = float(project_total)

    if not result["invoice_number"]:
        result["confidence"] -= 0.2
    if not result["line_items"]:
        result["confidence"] -= 0.3

    return result


def parse_xlsx(xlsx_path):
    """Parse an XLSX invoice file. Detects format by sheet name."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {"format": "error", "confidence": 0, "error": str(e), "line_items": []}

    ws = wb.active
    sheet_name = ws.title

    if sheet_name == "Invoice":
        result = parse_xlsx_professional(ws)
    elif sheet_name == "Simple Invoice":
        result = parse_xlsx_simple(ws)
    else:
        # Try to detect by checking cell B17 for "Task" header
        if ws["B17"].value and "Task" in str(ws["B17"].value):
            result = parse_xlsx_professional(ws)
        elif ws["A21"].value and "DESCRIPTION" in str(ws["A21"].value).upper():
            result = parse_xlsx_simple(ws)
        else:
            result = {"format": "unknown", "confidence": 0.3, "line_items": []}

    result["source_path"] = xlsx_path
    result["filename"] = os.path.basename(xlsx_path)
    result["source_type"] = "xlsx"
    result["confidence"] = max(0.0, min(1.0, result.get("confidence", 0)))
    return result


# ── PDF parsing (fallback) ────────────────────────────────────────────

def parse_money(s):
    if not s:
        return None
    cleaned = re.sub(r'[$\s,]', '', str(s).strip())
    if cleaned in ('-', ''):
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_pdf(pdf_path):
    """Parse a PDF invoice (fallback when no XLSX available)."""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        return {"format": "error", "confidence": 0, "error": str(e), "line_items": []}

    is_professional = any(kw in text for kw in ("% Complete", "Previous Fee", "Current Fee", "Professional Services"))

    if is_professional:
        result = _parse_pdf_professional(text)
    else:
        result = _parse_pdf_simple(text)

    result["source_path"] = pdf_path
    result["filename"] = os.path.basename(pdf_path)
    result["source_type"] = "pdf"
    result["confidence"] = max(0.0, min(1.0, result.get("confidence", 0)))
    return result


def _parse_pdf_professional(text):
    result = {"format": "professional", "confidence": 0.85, "line_items": []}

    inv_num = re.search(r'Invoice Number:\s*(\S+)', text)
    inv_date = re.search(r'Invoice Date:\s*(\S+)', text)
    if inv_num:
        result["invoice_number"] = inv_num.group(1)
    if inv_date:
        result["date"] = inv_date.group(1)

    # Task lines: TaskName $Fee Percent $PrevBilling $CurrentBilling
    task_pattern = re.compile(
        r'^(.+?)\s+\$?([\d,]+\.?\d*)\s+([\d.]+)%?\s+\$?([\d,]+\.?\d*)\s+\$?([\d,]+\.?\d*)\s*$',
        re.MULTILINE
    )
    for m in task_pattern.finditer(text):
        name = m.group(1).strip()
        if name.lower() in ('task', 'total', 'previous fee', 'current fee'):
            continue
        fee = parse_money(m.group(2))
        pct = float(m.group(3))
        prev = parse_money(m.group(4))
        current = parse_money(m.group(5))
        if fee and fee > 0:
            result["line_items"].append({
                "name": name,
                "unit_price": fee,
                "quantity": pct,
                "previous_billing": prev or 0,
                "amount": current or 0,
            })

    total_match = re.search(r'Invoice Total\s+\$?([\d,]+\.?\d*)', text)
    if total_match:
        result["total_due"] = parse_money(total_match.group(1))
    else:
        result["total_due"] = sum(li["amount"] for li in result["line_items"])

    if not result.get("invoice_number"):
        result["confidence"] -= 0.2
    if not result["line_items"]:
        result["confidence"] -= 0.3
    return result


def _parse_pdf_simple(text):
    result = {"format": "simple", "confidence": 0.80, "line_items": []}

    inv_num = re.search(r'INVOICE\s*#\s*(\S+)', text)
    if inv_num:
        result["invoice_number"] = inv_num.group(1)

    # Line items: name qty $rate $amount
    line_pattern = re.compile(
        r'^(.+?)\s+([\d.]+)\s+(\$[\s\d,]+\.?\d*)\s+(\$[\s\d,]+\.?\d*)\s*$',
        re.MULTILINE
    )
    for m in line_pattern.finditer(text):
        name = m.group(1).strip()
        if name.lower() in ('description', 'previous invoices') or 'quantity' in name.lower():
            continue
        qty = float(m.group(2))
        rate = parse_money(m.group(3))
        amount = parse_money(m.group(4))
        if amount and amount > 0:
            result["line_items"].append({
                "name": name,
                "quantity": qty,
                "unit_price": rate or 0,
                "amount": amount,
                "previous_billing": 0,
            })

    total_match = re.search(r'TOTAL\s+DUE\s+(\$[\s\d,]+\.?\d*)', text)
    if total_match:
        result["total_due"] = parse_money(total_match.group(1))
    else:
        result["total_due"] = sum(li["amount"] for li in result["line_items"])

    if not result.get("invoice_number"):
        result["confidence"] -= 0.2
    if not result["line_items"]:
        result["confidence"] -= 0.3
    return result


# ── Import to database ────────────────────────────────────────────────

def import_invoice(conn, db, inv, project_id, contract_id=None):
    """Import an invoice with line items directly into the database."""
    inv_id = generate_id("inv-")
    inv_number = inv.get("invoice_number", "")
    total_due = inv.get("total_due", 0) or 0
    inv_type = "task" if inv.get("format") == "professional" else "list"
    description = inv.get("description", "") or f"Imported from {inv['filename']}"
    now = datetime.now().isoformat()

    # Parse date from invoice
    date_str = inv.get("date")
    created_at = now
    if date_str:
        try:
            # Try various date formats
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    dt = datetime.strptime(date_str, fmt)
                    created_at = dt.isoformat()
                    break
                except ValueError:
                    continue
        except Exception:
            pass

    # Find previous invoice in chain for this project
    db.execute(
        "SELECT id FROM invoices WHERE project_id = %s AND deleted_at IS NULL "
        "ORDER BY created_at DESC LIMIT 1",
        (project_id,),
    )
    prev_invoice = db.fetchone()
    previous_invoice_id = prev_invoice["id"] if prev_invoice else None

    # Insert invoice
    db.execute(
        "INSERT INTO invoices (id, invoice_number, project_id, contract_id, previous_invoice_id, "
        "type, description, total_due, pdf_path, sent_status, paid_status, created_at, updated_at) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'sent', 'paid', %s, %s)",
        (inv_id, inv_number, project_id, contract_id, previous_invoice_id,
         inv_type, description, total_due, inv.get("source_path"), created_at, now),
    )

    # Insert line items
    for i, li in enumerate(inv.get("line_items", [])):
        li_id = generate_id("li-")
        db.execute(
            "INSERT INTO invoice_line_items (id, invoice_id, sort_order, name, description, "
            "quantity, unit_price, amount, previous_billing, created_at) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (li_id, inv_id, i + 1, li["name"], None,
             li.get("quantity", 1), li.get("unit_price", 0),
             li.get("amount", 0), li.get("previous_billing", 0), now),
        )

    return inv_id


# ── Main ──────────────────────────────────────────────────────────────

def main():
    do_import_all = "--import" in sys.argv
    do_import_high = "--import-high" in sys.argv

    # Get projects from API
    try:
        resp = requests.get(f"{API_BASE}/projects", timeout=3)
        resp.raise_for_status()
    except Exception as e:
        print(f"Server not reachable at {API_BASE}: {e}")
        sys.exit(1)

    projects = resp.json()
    # Get full details for each
    detailed = []
    for p in projects:
        detail = requests.get(f"{API_BASE}/projects/{p['id']}").json()
        detailed.append(detail)

    print(f"Found {len(detailed)} projects in database\n")

    all_invoices = []

    for proj in detailed:
        dropbox_path = proj.get("dropbox_path")
        data_path = os.path.join(TIE_BASE, dropbox_path) if dropbox_path else None
        if not data_path or not os.path.isdir(data_path):
            continue

        files = find_invoice_files(data_path)
        xlsx_files = files["xlsx"]
        pdf_files = files["pdf"]

        if not xlsx_files and not pdf_files:
            continue

        project_name = proj.get("project_name") or proj.get("name", "?")
        project_id = proj["id"]
        # Get contract_id if one exists
        contracts = proj.get("contracts", [])
        contract_id = contracts[0]["id"] if contracts else None

        print(f"\n{'─'*60}")
        print(f"  {project_name} ({project_id})")
        print(f"   XLSX: {len(xlsx_files)}, PDF: {len(pdf_files)}")

        # Track which invoice numbers we've already parsed from XLSX
        parsed_numbers = set()

        # Phase 1: Parse XLSX files (preferred)
        for xlsx_path in sorted(xlsx_files):
            parsed = parse_xlsx(xlsx_path)
            parsed["project_id"] = project_id
            parsed["project_name"] = project_name
            parsed["contract_id"] = contract_id

            inv_num = parsed.get("invoice_number", "?")
            total = parsed.get("total_due")
            li_count = len(parsed.get("line_items", []))
            total_str = f"${total:,.2f}" if total else "?"

            print(f"   [xlsx] {parsed['filename']}")
            print(f"          #{inv_num} | {total_str} | {li_count} items | {parsed['confidence']*100:.0f}%")
            for li in parsed.get("line_items", []):
                if parsed["format"] == "professional":
                    print(f"            - {li['name']}: fee=${li['unit_price']:,.0f}, {li['quantity']:.0f}%, prev=${li['previous_billing']:,.2f}, current=${li['amount']:,.2f}")
                else:
                    print(f"            - {li['name']}: {li['quantity']} x ${li['unit_price']:,.2f} = ${li['amount']:,.2f}")

            if inv_num:
                parsed_numbers.add(inv_num)
            all_invoices.append(parsed)

        # Phase 2: Parse PDF files (only if no matching XLSX)
        for pdf_path in sorted(pdf_files):
            parsed = parse_pdf(pdf_path)
            parsed["project_id"] = project_id
            parsed["project_name"] = project_name
            parsed["contract_id"] = contract_id

            inv_num = parsed.get("invoice_number", "?")

            # Skip if we already have this invoice from XLSX
            if inv_num in parsed_numbers:
                print(f"   [pdf]  {parsed['filename']} — skipped (have XLSX)")
                continue

            total = parsed.get("total_due")
            li_count = len(parsed.get("line_items", []))
            total_str = f"${total:,.2f}" if total else "?"

            print(f"   [pdf]  {parsed['filename']}")
            print(f"          #{inv_num} | {total_str} | {li_count} items | {parsed['confidence']*100:.0f}%")
            for li in parsed.get("line_items", []):
                print(f"            - {li['name']}: ${li['amount']:,.2f}")

            all_invoices.append(parsed)

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")

    total_invoices = len(all_invoices)
    from_xlsx = sum(1 for i in all_invoices if i.get("source_type") == "xlsx")
    from_pdf = sum(1 for i in all_invoices if i.get("source_type") == "pdf")
    with_items = sum(1 for i in all_invoices if i.get("line_items"))
    high_conf = [i for i in all_invoices if i["confidence"] >= 0.85]
    med_conf = [i for i in all_invoices if 0.5 <= i["confidence"] < 0.85]
    low_conf = [i for i in all_invoices if i["confidence"] < 0.5]
    total_amount = sum(i.get("total_due", 0) or 0 for i in all_invoices)

    print(f"Total invoices found: {total_invoices} (XLSX: {from_xlsx}, PDF: {from_pdf})")
    print(f"  With line items: {with_items}")
    print(f"  High confidence (>=85%): {len(high_conf)}")
    print(f"  Medium confidence (50-84%): {len(med_conf)}")
    print(f"  Low confidence (<50%): {len(low_conf)}")
    print(f"Total amount: ${total_amount:,.2f}")

    if med_conf or low_conf:
        print(f"\n{'─'*60}")
        print("  NEEDS MANUAL REVIEW:")
        print(f"{'─'*60}")
        for inv in med_conf + low_conf:
            total = inv.get('total_due')
            total_str = f"${total:,.2f}" if total else "NO TOTAL"
            print(f"  {inv['project_name']}: {inv['filename']} ({inv['confidence']*100:.0f}%) — {total_str}")

    # Write report
    report_path = "/mnt/d/repos/podium/scripts/invoice_extraction_report.json"
    with open(report_path, "w") as f:
        json.dump(all_invoices, f, indent=2, default=str)
    print(f"\nReport: {report_path}")

    # Import
    if do_import_all or do_import_high:
        min_conf = 0.85 if do_import_high else 0.3
        to_import = [i for i in all_invoices if i["confidence"] >= min_conf and (i.get("total_due") or 0) > 0]

        print(f"\n{'='*60}")
        print(f"IMPORTING {len(to_import)} invoices (min confidence {min_conf*100:.0f}%)")
        print(f"{'='*60}")

        conn = psycopg2.connect(DB_URL)
        conn.autocommit = False
        db = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        imported = 0
        # Sort by project then date for proper chaining
        to_import.sort(key=lambda x: (x.get("project_id", ""), x.get("date", ""), x.get("invoice_number", "")))

        for inv in to_import:
            try:
                inv_id = import_invoice(conn, db, inv, inv["project_id"], inv.get("contract_id"))
                conn.commit()
                li_count = len(inv.get("line_items", []))
                total_str = f"${inv.get('total_due', 0):,.2f}"
                print(f"  + {inv.get('invoice_number','?')}: {total_str} ({li_count} items) -> {inv_id} ({inv['project_name']})")
                imported += 1
            except Exception as e:
                conn.rollback()
                print(f"  ! {inv.get('invoice_number','?')} failed: {e}")

        conn.close()
        print(f"\nImported: {imported}")
    else:
        print(f"\nRun with --import-high to import {len(high_conf)} high-confidence invoices.")
        print(f"Run with --import to import all with total > $0.")


if __name__ == "__main__":
    main()
